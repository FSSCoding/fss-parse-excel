#!/usr/bin/env python3
"""
Cell Manager - Precise cell and range operations for Excel files
Handles A1 notation, formulas, and batch operations with safety.
"""

import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple, Union
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries, coordinate_to_tuple, get_column_letter
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string

from .converters import SafetyConfig, FileSafetyManager

class CellManager:
    """
    Manages precise cell and range operations with safety controls.
    Designed for CLI agents requiring exact cell-level control.
    """
    
    def __init__(self, file_path: Path, safety_config: SafetyConfig = None):
        self.file_path = Path(file_path)
        self.safety = FileSafetyManager(safety_config or SafetyConfig())
        self._workbook = None
        self._active_sheet = None
    
    def _ensure_workbook_loaded(self):
        """Ensure workbook is loaded and ready."""
        if self._workbook is None:
            if not self.file_path.exists():
                raise FileNotFoundError(f"Excel file not found: {self.file_path}")
            
            self._workbook = load_workbook(
                filename=str(self.file_path),
                read_only=False,
                data_only=False  # Keep formulas
            )
    
    def _parse_cell_reference(self, cell_ref: str) -> Tuple[str, str]:
        """
        Parse cell reference into sheet and cell parts.
        Supports: A1, Sheet1!A1, 'Sheet Name'!A1
        """
        if '!' in cell_ref:
            sheet_part, cell_part = cell_ref.rsplit('!', 1)
            # Remove quotes if present
            sheet_name = sheet_part.strip("'\"")
            return sheet_name, cell_part
        else:
            return None, cell_ref
    
    def _get_worksheet(self, sheet_name: Optional[str] = None):
        """Get worksheet by name or use active sheet."""
        self._ensure_workbook_loaded()
        
        if sheet_name:
            if sheet_name not in self._workbook.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' not found. Available: {self._workbook.sheetnames}")
            return self._workbook[sheet_name]
        else:
            return self._workbook.active
    
    def get_cell_value(self, cell_ref: str, sheet: Optional[str] = None) -> Any:
        """
        Get value from a specific cell.
        
        Args:
            cell_ref: Cell reference (A1, B2, etc.) or Sheet!Cell
            sheet: Sheet name (optional if specified in cell_ref)
            
        Returns:
            Cell value (can be string, number, datetime, formula, etc.)
        """
        try:
            sheet_name, cell_addr = self._parse_cell_reference(cell_ref)
            if sheet_name:
                sheet = sheet_name
            
            ws = self._get_worksheet(sheet)
            cell = ws[cell_addr]
            
            # Return the actual value (not formula result if data_only=False)
            return cell.value
            
        except Exception as e:
            raise ValueError(f"Error getting cell {cell_ref}: {e}")
    
    def get_cell_formula(self, cell_ref: str, sheet: Optional[str] = None) -> Optional[str]:
        """Get formula from a cell (returns None if no formula)."""
        try:
            sheet_name, cell_addr = self._parse_cell_reference(cell_ref)
            if sheet_name:
                sheet = sheet_name
            
            ws = self._get_worksheet(sheet)
            cell = ws[cell_addr]
            
            # openpyxl stores formulas in the value if data_only=False
            if hasattr(cell, 'data_type') and cell.data_type == 'f':
                return cell.value
            return None
            
        except Exception as e:
            raise ValueError(f"Error getting formula from {cell_ref}: {e}")
    
    def set_cell_value(self, cell_ref: str, value: Any, sheet: Optional[str] = None) -> bool:
        """
        Set value in a specific cell.
        
        Args:
            cell_ref: Cell reference (A1, B2, etc.) or Sheet!Cell
            value: Value to set (string, number, formula starting with =)
            sheet: Sheet name (optional if specified in cell_ref)
            
        Returns:
            True if successful, False otherwise
        """
        try:
            # Safety check
            can_proceed, reason = self.safety.safe_write_check(self.file_path, self.file_path)
            if not can_proceed:
                raise RuntimeError(f"Safety check failed: {reason}")
            
            sheet_name, cell_addr = self._parse_cell_reference(cell_ref)
            if sheet_name:
                sheet = sheet_name
            
            ws = self._get_worksheet(sheet)
            cell = ws[cell_addr]
            
            # Set the value (openpyxl handles formulas automatically if starting with =)
            cell.value = value
            
            # Save the file
            self._workbook.save(str(self.file_path))
            
            return True
            
        except Exception as e:
            print(f"❌ Error setting cell {cell_ref}: {e}")
            return False
    
    def get_range_values(self, range_ref: str, sheet: Optional[str] = None) -> List[List[Any]]:
        """
        Get values from a range of cells.
        
        Args:
            range_ref: Range reference (A1:C3, etc.) or Sheet!Range
            sheet: Sheet name (optional if specified in range_ref)
            
        Returns:
            2D list of cell values
        """
        try:
            sheet_name, range_addr = self._parse_cell_reference(range_ref)
            if sheet_name:
                sheet = sheet_name
            
            ws = self._get_worksheet(sheet)
            
            # Get the range
            cell_range = ws[range_addr]
            
            # Convert to 2D list
            if hasattr(cell_range, '__iter__') and not isinstance(cell_range, str):
                # Multiple cells
                result = []
                for row in cell_range:
                    if hasattr(row, '__iter__'):
                        # Multiple cells in row
                        result.append([cell.value for cell in row])
                    else:
                        # Single cell in row
                        result.append([row.value])
                return result
            else:
                # Single cell
                return [[cell_range.value]]
                
        except Exception as e:
            raise ValueError(f"Error getting range {range_ref}: {e}")
    
    def set_range_values(self, range_ref: str, values: Union[List[List], List, Any], 
                        sheet: Optional[str] = None) -> bool:
        """
        Set values in a range of cells.
        
        Args:
            range_ref: Range reference (A1:C3, etc.)
            values: Values to set (2D list, 1D list, or single value)
            sheet: Sheet name
            
        Returns:
            True if successful, False otherwise
        """
        try:
            # Safety check
            can_proceed, reason = self.safety.safe_write_check(self.file_path, self.file_path)
            if not can_proceed:
                raise RuntimeError(f"Safety check failed: {reason}")
            
            sheet_name, range_addr = self._parse_cell_reference(range_ref)
            if sheet_name:
                sheet = sheet_name
            
            ws = self._get_worksheet(sheet)
            
            # Parse range boundaries
            min_col, min_row, max_col, max_row = range_boundaries(range_addr)
            
            # Ensure values is a 2D list
            if not isinstance(values, list):
                # Single value - fill entire range
                values = [[values for _ in range(max_col - min_col + 1)] 
                         for _ in range(max_row - min_row + 1)]
            elif values and not isinstance(values[0], list):
                # 1D list - convert to 2D
                if len(values) == (max_row - min_row + 1):
                    # Column vector
                    values = [[v] for v in values]
                elif len(values) == (max_col - min_col + 1):
                    # Row vector
                    values = [values]
                else:
                    # Fill first row, then repeat
                    values = [values]
            
            # Set values
            for row_idx, row_values in enumerate(values):
                if row_idx > (max_row - min_row):
                    break
                    
                for col_idx, value in enumerate(row_values):
                    if col_idx > (max_col - min_col):
                        break
                    
                    cell_row = min_row + row_idx
                    cell_col = min_col + col_idx
                    
                    ws.cell(row=cell_row, column=cell_col, value=value)
            
            # Save the file
            self._workbook.save(str(self.file_path))
            
            return True
            
        except Exception as e:
            print(f"❌ Error setting range {range_ref}: {e}")
            return False
    
    def copy_range(self, source_range: str, dest_cell: str, 
                   source_sheet: Optional[str] = None, dest_sheet: Optional[str] = None) -> bool:
        """Copy a range to another location."""
        try:
            # Get source values
            source_values = self.get_range_values(source_range, source_sheet)
            
            # Calculate destination range
            source_sheet_name, source_addr = self._parse_cell_reference(source_range)
            dest_sheet_name, dest_addr = self._parse_cell_reference(dest_cell)
            
            # Parse source range dimensions
            min_col, min_row, max_col, max_row = range_boundaries(source_addr)
            rows = max_row - min_row + 1
            cols = max_col - min_col + 1
            
            # Parse destination start
            dest_col_str, dest_row = coordinate_from_string(dest_addr)
            dest_col = column_index_from_string(dest_col_str)
            
            # Create destination range
            dest_end_col = get_column_letter(dest_col + cols - 1)
            dest_end_row = dest_row + rows - 1
            dest_range = f"{dest_addr}:{dest_end_col}{dest_end_row}"
            
            # Set destination values
            return self.set_range_values(
                dest_range, source_values, 
                dest_sheet or dest_sheet_name
            )
            
        except Exception as e:
            print(f"❌ Error copying range: {e}")
            return False
    
    def clear_range(self, range_ref: str, sheet: Optional[str] = None) -> bool:
        """Clear values from a range."""
        try:
            sheet_name, range_addr = self._parse_cell_reference(range_ref)
            if sheet_name:
                sheet = sheet_name
            
            ws = self._get_worksheet(sheet)
            
            # Clear the range
            for row in ws[range_addr]:
                if hasattr(row, '__iter__'):
                    for cell in row:
                        cell.value = None
                else:
                    row.value = None
            
            # Save the file
            self._workbook.save(str(self.file_path))
            
            return True
            
        except Exception as e:
            print(f"❌ Error clearing range {range_ref}: {e}")
            return False
    
    def find_cells(self, value: Any, sheet: Optional[str] = None, 
                   match_case: bool = False) -> List[str]:
        """
        Find all cells containing a specific value.
        
        Returns:
            List of cell references (e.g., ['A1', 'B5', 'C10'])
        """
        try:
            ws = self._get_worksheet(sheet)
            matches = []
            
            for row in ws.iter_rows():
                for cell in row:
                    cell_value = cell.value
                    
                    # Handle different comparison types
                    if cell_value is None:
                        continue
                    
                    if isinstance(value, str) and isinstance(cell_value, str):
                        if match_case:
                            match = cell_value == value
                        else:
                            match = cell_value.lower() == value.lower()
                    else:
                        match = cell_value == value
                    
                    if match:
                        matches.append(cell.coordinate)
            
            return matches
            
        except Exception as e:
            raise ValueError(f"Error finding cells: {e}")
    
    def get_cell_info(self, cell_ref: str, sheet: Optional[str] = None) -> Dict[str, Any]:
        """
        Get comprehensive information about a cell.
        
        Returns:
            Dictionary with value, formula, formatting, etc.
        """
        try:
            sheet_name, cell_addr = self._parse_cell_reference(cell_ref)
            if sheet_name:
                sheet = sheet_name
            
            ws = self._get_worksheet(sheet)
            cell = ws[cell_addr]
            
            info = {
                'coordinate': cell.coordinate,
                'value': cell.value,
                'data_type': getattr(cell, 'data_type', None),
                'formula': self.get_cell_formula(cell_ref, sheet),
                'number_format': cell.number_format,
                'font': {
                    'name': cell.font.name,
                    'size': cell.font.size,
                    'bold': cell.font.bold,
                    'italic': cell.font.italic,
                } if cell.font else None,
                'fill': {
                    'color': str(cell.fill.start_color.rgb) if cell.fill and cell.fill.start_color else None
                } if cell.fill else None,
                'alignment': {
                    'horizontal': cell.alignment.horizontal,
                    'vertical': cell.alignment.vertical
                } if cell.alignment else None,
                'hyperlink': cell.hyperlink.target if cell.hyperlink else None,
                'comment': cell.comment.text if cell.comment else None
            }
            
            return info
            
        except Exception as e:
            raise ValueError(f"Error getting cell info for {cell_ref}: {e}")
    
    def close(self):
        """Close the workbook and release resources."""
        if self._workbook:
            self._workbook.close()
            self._workbook = None