#!/usr/bin/env python3
"""
Excel - Global Spreadsheet Conversion Tool
Safe, hash-validated bidirectional conversion between .xlsx/.csv and structured formats.

Author: System Blueprint
Safety Features: Hash checking, collision detection, confirmation prompts
Supports: .xlsx, .xls, .csv, .tsv, .json, .yaml, .md (tables)
"""

import argparse
import hashlib
import json
import re
import sys
import yaml
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple, Union
from dataclasses import dataclass, asdict
import tempfile
import shutil
from datetime import datetime
import csv

try:
    import pandas as pd
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError:
    print("❌ Error: Required packages not installed. Run: pip install pandas openpyxl")
    sys.exit(1)

try:
    import xlrd  # For legacy .xls files
except ImportError:
    print("⚠️  Warning: xlrd not installed. Legacy .xls support disabled. Install with: pip install xlrd")
    xlrd = None


@dataclass
class SafetyConfig:
    """Configuration for safety mechanisms."""
    require_confirmation: bool = True
    create_backup: bool = True
    check_hash: bool = True
    prevent_overwrite: bool = True
    backup_suffix: str = ".backup"


@dataclass
class SheetMetadata:
    """Stores sheet-level formatting and structure information."""
    name: str = ""
    cell_formats: Dict[str, Dict[str, Any]] = None
    column_widths: Dict[str, float] = None
    row_heights: Dict[str, float] = None
    merged_cells: List[str] = None
    conditional_formats: List[Dict[str, Any]] = None
    data_validation: Dict[str, Any] = None
    charts: List[Dict[str, Any]] = None
    tables: List[Dict[str, Any]] = None
    formulas: Dict[str, str] = None
    hyperlinks: Dict[str, str] = None
    comments: Dict[str, str] = None
    
    def __post_init__(self):
        if self.cell_formats is None:
            self.cell_formats = {}
        if self.column_widths is None:
            self.column_widths = {}
        if self.row_heights is None:
            self.row_heights = {}
        if self.merged_cells is None:
            self.merged_cells = []
        if self.conditional_formats is None:
            self.conditional_formats = []
        if self.data_validation is None:
            self.data_validation = {}
        if self.charts is None:
            self.charts = []
        if self.tables is None:
            self.tables = []
        if self.formulas is None:
            self.formulas = {}
        if self.hyperlinks is None:
            self.hyperlinks = {}
        if self.comments is None:
            self.comments = {}


@dataclass
class WorkbookMetadata:
    """Stores workbook-level formatting and structure information."""
    sheets: Dict[str, SheetMetadata] = None
    defined_names: Dict[str, str] = None
    custom_properties: Dict[str, Any] = None
    file_hash: str = ""
    conversion_timestamp: str = ""
    original_format: str = ""
    
    def __post_init__(self):
        if self.sheets is None:
            self.sheets = {}
        if self.defined_names is None:
            self.defined_names = {}
        if self.custom_properties is None:
            self.custom_properties = {}


@dataclass
class ConversionConfig:
    """Configuration for Excel conversion operations."""
    # Output format settings
    preserve_formatting: bool = True
    preserve_formulas: bool = True
    preserve_charts: bool = False  # Charts are complex to convert
    preserve_comments: bool = True
    preserve_hyperlinks: bool = True
    
    # CSV settings
    csv_delimiter: str = ","
    csv_quotechar: str = '"'
    csv_encoding: str = "utf-8"
    csv_include_index: bool = False
    csv_include_header: bool = True
    
    # JSON settings
    json_orient: str = "records"  # records, index, values, columns
    json_indent: int = 2
    json_include_metadata: bool = True
    
    # YAML settings
    yaml_default_flow_style: bool = False
    yaml_include_metadata: bool = True
    
    # Markdown table settings
    md_table_alignment: str = "left"  # left, center, right
    md_max_col_width: int = 50
    md_include_sheet_names: bool = True
    
    # Excel output settings
    excel_engine: str = "openpyxl"  # openpyxl, xlsxwriter
    excel_index: bool = False
    excel_header: bool = True
    excel_freeze_panes: bool = True
    excel_autofilter: bool = True
    excel_table_style: str = "TableStyleMedium2"
    
    # Data processing
    skip_empty_rows: bool = True
    skip_empty_columns: bool = True
    trim_whitespace: bool = True
    convert_numeric: bool = True
    date_format: str = "%Y-%m-%d"
    datetime_format: str = "%Y-%m-%d %H:%M:%S"
    
    # Multi-sheet handling
    sheet_selection: Optional[List[str]] = None  # None = all sheets
    combine_sheets: bool = False
    sheet_name_column: str = "sheet_name"


class FileSafetyManager:
    """Handles file safety operations: hashing, collision detection, backups."""
    
    def __init__(self, safety_config: SafetyConfig = None):
        self.config = safety_config or SafetyConfig()
    
    def calculate_file_hash(self, file_path: Path) -> str:
        """Calculate SHA256 hash of file."""
        if not file_path.exists():
            return ""
        
        sha256_hash = hashlib.sha256()
        with open(file_path, "rb") as f:
            for byte_block in iter(lambda: f.read(4096), b""):
                sha256_hash.update(byte_block)
        return sha256_hash.hexdigest()
    
    def detect_conversion_collision(self, source_file: Path, target_file: Path) -> bool:
        """Check if target file would create a conversion collision."""
        if not target_file.exists():
            return False
        
        if source_file.stem == target_file.stem:
            source_hash = self.calculate_file_hash(source_file)
            target_hash = self.calculate_file_hash(target_file)
            return source_hash != target_hash
        
        return False
    
    def create_backup(self, file_path: Path) -> Optional[Path]:
        """Create backup of existing file."""
        if not file_path.exists():
            return None
        
        backup_path = file_path.with_suffix(f"{file_path.suffix}{self.config.backup_suffix}")
        counter = 1
        
        while backup_path.exists():
            backup_path = file_path.with_suffix(f"{file_path.suffix}{self.config.backup_suffix}.{counter}")
            counter += 1
        
        try:
            shutil.copy2(file_path, backup_path)
            return backup_path
        except Exception as e:
            print(f"⚠️  Warning: Could not create backup: {e}")
            return None
    
    def confirm_overwrite(self, file_path: Path) -> bool:
        """Get user confirmation for file overwrite."""
        if not self.config.require_confirmation:
            return True
        
        response = input(f"⚠️  File '{file_path}' exists. Overwrite? [y/N]: ").lower().strip()
        return response in ['y', 'yes']
    
    def safe_write_check(self, source_file: Path, target_file: Path) -> Tuple[bool, str]:
        """
        Comprehensive safety check before writing.
        Returns (can_proceed, reason)
        """
        if self.detect_conversion_collision(source_file, target_file):
            return False, f"Collision detected: {target_file} exists with different content"
        
        if target_file.exists():
            if self.config.prevent_overwrite:
                if not self.confirm_overwrite(target_file):
                    return False, "User cancelled overwrite"
            
            if self.config.create_backup:
                backup_path = self.create_backup(target_file)
                if backup_path:
                    print(f"✅ Backup created: {backup_path}")
        
        return True, "Safe to proceed"


class ExcelReader:
    """Reads Excel files and extracts data with metadata preservation."""
    
    def __init__(self, config: ConversionConfig = None):
        self.config = config or ConversionConfig()
        self.metadata = WorkbookMetadata()
    
    def read_file(self, file_path: str) -> Tuple[Dict[str, pd.DataFrame], WorkbookMetadata]:
        """
        Read Excel file and return data with metadata.
        Returns: (sheets_data, metadata)
        """
        path = Path(file_path)
        
        if not path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")
        
        # Set original format
        self.metadata.original_format = path.suffix.lower()
        
        # Route to appropriate reader
        if path.suffix.lower() == '.csv':
            return self._read_csv(file_path)
        elif path.suffix.lower() == '.tsv':
            return self._read_tsv(file_path)
        elif path.suffix.lower() in ['.xlsx', '.xlsm']:
            return self._read_xlsx(file_path)
        elif path.suffix.lower() == '.xls' and xlrd:
            return self._read_xls(file_path)
        else:
            raise ValueError(f"Unsupported file format: {path.suffix}")
    
    def _read_csv(self, file_path: str) -> Tuple[Dict[str, pd.DataFrame], WorkbookMetadata]:
        """Read CSV file with configurable options."""
        try:
            # Read CSV with pandas
            df = pd.read_csv(
                file_path,
                delimiter=self.config.csv_delimiter,
                quotechar=self.config.csv_quotechar,
                encoding=self.config.csv_encoding,
                header=0 if self.config.csv_include_header else None,
                index_col=0 if self.config.csv_include_index else None
            )
            
            # Apply data processing
            if self.config.skip_empty_rows:
                df = df.dropna(how='all')
            if self.config.skip_empty_columns:
                df = df.dropna(axis=1, how='all')
            if self.config.trim_whitespace:
                for col in df.select_dtypes(include=['object']):
                    df[col] = df[col].astype(str).str.strip()
            if self.config.convert_numeric:
                df = df.apply(pd.to_numeric, errors='ignore')
            
            # Create metadata
            metadata = WorkbookMetadata()
            metadata.original_format = '.csv'
            
            sheet_metadata = SheetMetadata(name='Sheet1')
            metadata.sheets['Sheet1'] = sheet_metadata
            
            return {'Sheet1': df}, metadata
            
        except Exception as e:
            raise ValueError(f"Error reading CSV file {file_path}: {e}")
    
    def _read_tsv(self, file_path: str) -> Tuple[Dict[str, pd.DataFrame], WorkbookMetadata]:
        """Read TSV file (Tab-Separated Values)."""
        # Temporarily override delimiter for TSV
        original_delimiter = self.config.csv_delimiter
        self.config.csv_delimiter = '\t'
        
        try:
            result = self._read_csv(file_path)
            # Update metadata to reflect TSV format
            result[1].original_format = '.tsv'
            return result
        finally:
            # Restore original delimiter
            self.config.csv_delimiter = original_delimiter
    
    def _read_xlsx(self, file_path: str) -> Tuple[Dict[str, pd.DataFrame], WorkbookMetadata]:
        """Read XLSX file with full metadata extraction."""
        from openpyxl import load_workbook
        from openpyxl.utils.dataframe import dataframe_to_rows
        import warnings
        
        # Load workbook with comprehensive features
        with warnings.catch_warnings():
            warnings.filterwarnings("ignore", message="Unknown extension is not supported and will be removed")
            wb = load_workbook(
                filename=file_path,
                read_only=False,
                data_only=not self.config.preserve_formulas,
                keep_vba=True
            )
        
        sheets_data = {}
        workbook_metadata = WorkbookMetadata()
        workbook_metadata.original_format = Path(file_path).suffix.lower()
        
        # Extract workbook-level metadata
        workbook_metadata.defined_names = {dn.name: dn.value for dn in wb.defined_names.definedName}
        workbook_metadata.custom_properties = dict(wb.custom_doc_props) if hasattr(wb, 'custom_doc_props') else {}
        
        # Process each sheet
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            sheet_metadata = SheetMetadata(name=sheet_name)
            
            # Extract data as DataFrame
            data_rows = []
            max_row = sheet.max_row or 0
            max_col = sheet.max_column or 0
            
            # Get column headers
            if max_row > 0 and self.config.csv_include_header:
                headers = []
                for col in range(1, max_col + 1):
                    cell = sheet.cell(row=1, column=col)
                    headers.append(str(cell.value) if cell.value is not None else f"Column_{col}")
                
                # Get data rows (skip header if present)
                start_row = 2 if self.config.csv_include_header else 1
                for row in range(start_row, max_row + 1):
                    row_data = []
                    for col in range(1, max_col + 1):
                        cell = sheet.cell(row=row, column=col)
                        
                        # Extract cell value and metadata
                        value = cell.value
                        if value is None and not self.config.preserve_formatting:
                            value = ""
                        
                        # Store cell formatting if requested
                        if self.config.preserve_formatting and cell.coordinate:
                            cell_format = {
                                'font': {
                                    'name': cell.font.name,
                                    'size': cell.font.size,
                                    'bold': cell.font.bold,
                                    'italic': cell.font.italic,
                                    'color': str(cell.font.color.rgb) if cell.font.color else None
                                } if cell.font else {},
                                'fill': {
                                    'color': str(cell.fill.start_color.rgb) if cell.fill and cell.fill.start_color else None
                                } if cell.fill else {},
                                'alignment': {
                                    'horizontal': cell.alignment.horizontal,
                                    'vertical': cell.alignment.vertical
                                } if cell.alignment else {}
                            }
                            sheet_metadata.cell_formats[cell.coordinate] = cell_format
                        
                        # Store formulas if requested
                        if self.config.preserve_formulas and hasattr(cell, 'formula') and cell.formula:
                            sheet_metadata.formulas[cell.coordinate] = cell.formula
                        
                        # Store hyperlinks if requested
                        if self.config.preserve_hyperlinks and cell.hyperlink:
                            sheet_metadata.hyperlinks[cell.coordinate] = cell.hyperlink.target
                        
                        # Store comments if requested
                        if self.config.preserve_comments and cell.comment:
                            sheet_metadata.comments[cell.coordinate] = cell.comment.text
                        
                        row_data.append(value)
                    
                    if not self.config.skip_empty_rows or any(cell != "" for cell in row_data):
                        data_rows.append(row_data)
            
            # Create DataFrame
            if data_rows:
                if self.config.csv_include_header and headers:
                    df = pd.DataFrame(data_rows, columns=headers)
                else:
                    df = pd.DataFrame(data_rows)
                
                # Clean up DataFrame
                if self.config.skip_empty_columns:
                    df = df.dropna(axis=1, how='all')
                if self.config.trim_whitespace:
                    for col in df.select_dtypes(include=['object']):
                        df[col] = df[col].astype(str).str.strip()
                if self.config.convert_numeric:
                    df = df.apply(pd.to_numeric, errors='ignore')
            else:
                df = pd.DataFrame()
            
            # Extract column widths and row heights
            if self.config.preserve_formatting:
                for col_letter, dimension in sheet.column_dimensions.items():
                    if dimension.width:
                        sheet_metadata.column_widths[col_letter] = dimension.width
                
                for row_num, dimension in sheet.row_dimensions.items():
                    if dimension.height:
                        sheet_metadata.row_heights[str(row_num)] = dimension.height
            
            # Extract merged cells
            sheet_metadata.merged_cells = [str(merged_range) for merged_range in sheet.merged_cells.ranges]
            
            # Extract tables
            if hasattr(sheet, 'tables'):
                for table in sheet.tables.values():
                    table_info = {
                        'name': table.name,
                        'range': str(table.ref),
                        'style': table.tableStyleInfo.name if table.tableStyleInfo else None
                    }
                    sheet_metadata.tables.append(table_info)
            
            sheets_data[sheet_name] = df
            workbook_metadata.sheets[sheet_name] = sheet_metadata
        
        wb.close()
        return sheets_data, workbook_metadata
    
    def _read_xls(self, file_path: str) -> Tuple[Dict[str, pd.DataFrame], WorkbookMetadata]:
        """Read legacy XLS file with xlrd."""
        if not xlrd:
            raise ImportError("xlrd required for XLS parsing. Install with: pip install xlrd")
        
        try:
            # Open XLS workbook
            workbook = xlrd.open_workbook(str(file_path))
            if not workbook.sheet_names():
                raise ValueError(f"No sheets found in XLS file: {file_path}")
            
            sheets_data = {}
            workbook_metadata = WorkbookMetadata()
            workbook_metadata.original_format = '.xls'
            
            # Process each sheet
            for sheet_index, sheet_name in enumerate(workbook.sheet_names()):
                sheet = workbook.sheet_by_index(sheet_index)
                sheet_metadata = SheetMetadata(name=sheet_name)
                
                if sheet.nrows == 0 or sheet.ncols == 0:
                    sheets_data[sheet_name] = pd.DataFrame()
                    workbook_metadata.sheets[sheet_name] = sheet_metadata
                    continue
                
                # Extract data with enterprise-grade processing
                data_rows = []
                max_rows = min(sheet.nrows, 10000)  # Safety limit
                max_cols = min(sheet.ncols, 256)    # Excel XLS limit
                
                # Get headers if requested
                headers = None
                start_row = 0
                if self.config.csv_include_header and sheet.nrows > 0:
                    headers = []
                    for col in range(max_cols):
                        try:
                            cell_value = sheet.cell_value(0, col)
                            headers.append(str(cell_value) if cell_value else f"Column_{col+1}")
                        except Exception:
                            headers.append(f"Column_{col+1}")
                    start_row = 1
                
                # Extract data rows
                for row_index in range(start_row, max_rows):
                    row_data = []
                    has_data = False
                    
                    for col_index in range(max_cols):
                        try:
                            cell_value = sheet.cell_value(row_index, col_index)
                            cell_type = sheet.cell_type(row_index, col_index)
                            
                            # Format cell value based on type
                            if cell_type == xlrd.XL_CELL_EMPTY:
                                formatted_value = ""
                            elif cell_type == xlrd.XL_CELL_TEXT:
                                formatted_value = str(cell_value).strip()
                            elif cell_type == xlrd.XL_CELL_NUMBER:
                                # Handle integers vs floats
                                if isinstance(cell_value, float) and cell_value.is_integer():
                                    formatted_value = int(cell_value)
                                else:
                                    formatted_value = cell_value
                            elif cell_type == xlrd.XL_CELL_DATE:
                                try:
                                    # Convert Excel date to datetime
                                    date_tuple = xlrd.xldate_as_tuple(cell_value, workbook.datemode)
                                    if date_tuple[:3] != (0, 0, 0):  # Valid date
                                        formatted_value = datetime(*date_tuple).strftime(self.config.date_format)
                                    else:
                                        formatted_value = cell_value
                                except Exception:
                                    formatted_value = cell_value
                            elif cell_type == xlrd.XL_CELL_BOOLEAN:
                                formatted_value = bool(cell_value)
                            elif cell_type == xlrd.XL_CELL_ERROR:
                                formatted_value = "#ERROR"
                            elif cell_type == xlrd.XL_CELL_FORMULA:
                                formatted_value = cell_value  # Formula result
                                if self.config.preserve_formulas:
                                    # Note: xlrd doesn't provide formula text directly
                                    sheet_metadata.formulas[f"{row_index+1}:{col_index+1}"] = "FORMULA_PRESENT"
                            else:
                                formatted_value = cell_value
                            
                            row_data.append(formatted_value)
                            
                            if formatted_value != "":
                                has_data = True
                                
                        except Exception as e:
                            # Handle cell read errors gracefully
                            row_data.append("")
                    
                    # Skip empty rows if configured
                    if self.config.skip_empty_rows and not has_data:
                        continue
                        
                    data_rows.append(row_data)
                
                # Create DataFrame
                if data_rows:
                    if headers:
                        df = pd.DataFrame(data_rows, columns=headers[:len(data_rows[0])])
                    else:
                        df = pd.DataFrame(data_rows)
                    
                    # Apply data processing
                    if self.config.skip_empty_columns:
                        df = df.dropna(axis=1, how='all')
                    if self.config.trim_whitespace:
                        for col in df.select_dtypes(include=['object']):
                            df[col] = df[col].astype(str).str.strip()
                    if self.config.convert_numeric:
                        df = df.apply(pd.to_numeric, errors='ignore')
                else:
                    df = pd.DataFrame()
                
                sheets_data[sheet_name] = df
                workbook_metadata.sheets[sheet_name] = sheet_metadata
            
            return sheets_data, workbook_metadata
            
        except Exception as e:
            raise ValueError(f"Error reading XLS file {file_path}: {e}")


class ExcelWriter:
    """Writes data to various Excel and tabular formats."""
    
    def __init__(self, config: ConversionConfig = None):
        self.config = config or ConversionConfig()
    
    def write_file(self, sheets_data: Dict[str, pd.DataFrame], 
                   file_path: str, metadata: WorkbookMetadata = None) -> bool:
        """
        Write data to file in appropriate format.
        Returns: success status
        """
        path = Path(file_path)
        
        # Route to appropriate writer
        if path.suffix.lower() == '.csv':
            return self._write_csv(sheets_data, file_path, metadata)
        elif path.suffix.lower() == '.tsv':
            return self._write_tsv(sheets_data, file_path, metadata)
        elif path.suffix.lower() in ['.xlsx', '.xlsm']:
            return self._write_xlsx(sheets_data, file_path, metadata)
        elif path.suffix.lower() == '.json':
            return self._write_json(sheets_data, file_path, metadata)
        elif path.suffix.lower() in ['.yml', '.yaml']:
            return self._write_yaml(sheets_data, file_path, metadata)
        elif path.suffix.lower() == '.md':
            return self._write_markdown(sheets_data, file_path, metadata)
        else:
            raise ValueError(f"Unsupported output format: {path.suffix}")
    
    def _write_csv(self, sheets_data: Dict[str, pd.DataFrame], 
                   file_path: str, metadata: WorkbookMetadata = None) -> bool:
        """Write to CSV format."""
        try:
            # Handle multi-sheet data
            if self.config.combine_sheets and len(sheets_data) > 1:
                # Combine all sheets into one CSV
                combined_data = []
                for sheet_name, df in sheets_data.items():
                    if self.config.sheet_name_column:
                        df = df.copy()
                        df[self.config.sheet_name_column] = sheet_name
                    combined_data.append(df)
                
                combined_df = pd.concat(combined_data, ignore_index=True)
                
                # Write combined CSV
                combined_df.to_csv(
                    file_path,
                    sep=self.config.csv_delimiter,
                    quotechar=self.config.csv_quotechar,
                    encoding=self.config.csv_encoding,
                    index=self.config.csv_include_index,
                    header=self.config.csv_include_header
                )
                
            elif len(sheets_data) == 1:
                # Single sheet - write directly
                df = next(iter(sheets_data.values()))
                df.to_csv(
                    file_path,
                    sep=self.config.csv_delimiter,
                    quotechar=self.config.csv_quotechar,
                    encoding=self.config.csv_encoding,
                    index=self.config.csv_include_index,
                    header=self.config.csv_include_header
                )
                
            else:
                # Multiple sheets - create separate files
                base_path = Path(file_path)
                for sheet_name, df in sheets_data.items():
                    sheet_file = base_path.parent / f"{base_path.stem}_{sheet_name}{base_path.suffix}"
                    df.to_csv(
                        sheet_file,
                        sep=self.config.csv_delimiter,
                        quotechar=self.config.csv_quotechar,
                        encoding=self.config.csv_encoding,
                        index=self.config.csv_include_index,
                        header=self.config.csv_include_header
                    )
            
            # Write metadata as companion file if requested
            if metadata and self.config.preserve_formatting:
                metadata_file = Path(file_path).with_suffix('.metadata.json')
                with open(metadata_file, 'w', encoding='utf-8') as f:
                    json.dump(asdict(metadata), f, indent=2, default=str)
            
            return True
            
        except Exception as e:
            print(f"❌ Error writing CSV: {e}")
            return False
    
    def _write_tsv(self, sheets_data: Dict[str, pd.DataFrame], 
                   file_path: str, metadata: WorkbookMetadata = None) -> bool:
        """Write to TSV format (Tab-Separated Values)."""
        # Temporarily override delimiter for TSV
        original_delimiter = self.config.csv_delimiter
        self.config.csv_delimiter = '\t'
        
        try:
            return self._write_csv(sheets_data, file_path, metadata)
        finally:
            # Restore original delimiter
            self.config.csv_delimiter = original_delimiter
    
    def _write_xlsx(self, sheets_data: Dict[str, pd.DataFrame], 
                    file_path: str, metadata: WorkbookMetadata = None) -> bool:
        """Write to XLSX format with full formatting restoration."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl.worksheet.table import Table, TableStyleInfo
        
        try:
            wb = Workbook()
            
            # Remove default sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            
            # Process each sheet
            for sheet_name, df in sheets_data.items():
                ws = wb.create_sheet(title=sheet_name)
                
                # Get sheet metadata if available
                sheet_metadata = None
                if metadata and sheet_name in metadata.sheets:
                    sheet_metadata = metadata.sheets[sheet_name]
                
                # Add data to worksheet
                if not df.empty:
                    # Write headers if requested
                    if self.config.excel_header:
                        for col_num, column_title in enumerate(df.columns, 1):
                            cell = ws.cell(row=1, column=col_num, value=column_title)
                            # Make headers bold
                            cell.font = Font(bold=True)
                        start_row = 2
                    else:
                        start_row = 1
                    
                    # Write data
                    for r_idx, row in enumerate(dataframe_to_rows(df, index=self.config.excel_index, header=False), start_row):
                        for c_idx, value in enumerate(row, 1):
                            cell = ws.cell(row=r_idx, column=c_idx, value=value)
                            
                            # Restore cell formatting if available
                            if sheet_metadata and cell.coordinate in sheet_metadata.cell_formats:
                                cell_format = sheet_metadata.cell_formats[cell.coordinate]
                                
                                # Apply font formatting
                                if 'font' in cell_format and cell_format['font']:
                                    font_info = cell_format['font']
                                    cell.font = Font(
                                        name=font_info.get('name', 'Calibri'),
                                        size=font_info.get('size', 11),
                                        bold=font_info.get('bold', False),
                                        italic=font_info.get('italic', False),
                                        color=font_info.get('color')
                                    )
                                
                                # Apply fill formatting
                                if 'fill' in cell_format and cell_format['fill'].get('color'):
                                    cell.fill = PatternFill(
                                        start_color=cell_format['fill']['color'],
                                        end_color=cell_format['fill']['color'],
                                        fill_type='solid'
                                    )
                                
                                # Apply alignment
                                if 'alignment' in cell_format and cell_format['alignment']:
                                    align_info = cell_format['alignment']
                                    cell.alignment = Alignment(
                                        horizontal=align_info.get('horizontal'),
                                        vertical=align_info.get('vertical')
                                    )
                    
                    # Restore formulas if available
                    if sheet_metadata and sheet_metadata.formulas:
                        for cell_coord, formula in sheet_metadata.formulas.items():
                            try:
                                if formula != "FORMULA_PRESENT":  # Skip XLS placeholders
                                    ws[cell_coord].value = formula
                            except Exception as e:
                                print(f"⚠️  Warning: Could not restore formula at {cell_coord}: {e}")
                    
                    # Restore hyperlinks if available
                    if sheet_metadata and sheet_metadata.hyperlinks:
                        for cell_coord, target in sheet_metadata.hyperlinks.items():
                            try:
                                ws[cell_coord].hyperlink = target
                                ws[cell_coord].style = "Hyperlink"
                            except Exception as e:
                                print(f"⚠️  Warning: Could not restore hyperlink at {cell_coord}: {e}")
                    
                    # Restore column widths if available
                    if sheet_metadata and sheet_metadata.column_widths:
                        for col_letter, width in sheet_metadata.column_widths.items():
                            try:
                                ws.column_dimensions[col_letter].width = width
                            except Exception as e:
                                print(f"⚠️  Warning: Could not restore column width for {col_letter}: {e}")
                    
                    # Restore row heights if available
                    if sheet_metadata and sheet_metadata.row_heights:
                        for row_num, height in sheet_metadata.row_heights.items():
                            try:
                                ws.row_dimensions[int(row_num)].height = height
                            except Exception as e:
                                print(f"⚠️  Warning: Could not restore row height for row {row_num}: {e}")
                    
                    # Restore merged cells if available
                    if sheet_metadata and sheet_metadata.merged_cells:
                        for merge_range in sheet_metadata.merged_cells:
                            try:
                                ws.merge_cells(merge_range)
                            except Exception as e:
                                print(f"⚠️  Warning: Could not restore merged cells {merge_range}: {e}")
                    
                    # Add Excel table if requested
                    if self.config.excel_autofilter and not df.empty:
                        try:
                            # Create table range
                            max_row = ws.max_row
                            max_col = ws.max_column
                            table_range = f"A1:{ws.cell(max_row, max_col).coordinate}"
                            
                            # Create table
                            table = Table(displayName=f"Table_{sheet_name}", ref=table_range)
                            
                            # Add table style
                            style = TableStyleInfo(
                                name=self.config.excel_table_style,
                                showFirstColumn=False,
                                showLastColumn=False,
                                showRowStripes=True,
                                showColumnStripes=False
                            )
                            table.tableStyleInfo = style
                            
                            ws.add_table(table)
                        except Exception as e:
                            print(f"⚠️  Warning: Could not create table for {sheet_name}: {e}")
                            # Fallback to auto filter
                            try:
                                ws.auto_filter.ref = f"A1:{ws.cell(ws.max_row, ws.max_column).coordinate}"
                            except Exception:
                                pass
                    
                    # Freeze panes if requested
                    if self.config.excel_freeze_panes and self.config.excel_header:
                        try:
                            ws.freeze_panes = "A2"  # Freeze header row
                        except Exception as e:
                            print(f"⚠️  Warning: Could not freeze panes for {sheet_name}: {e}")
            
            # Restore workbook-level defined names if available
            if metadata and metadata.defined_names:
                for name, value in metadata.defined_names.items():
                    try:
                        wb.create_named_range(name, ws, value)
                    except Exception as e:
                        print(f"⚠️  Warning: Could not restore named range {name}: {e}")
            
            # Save workbook
            wb.save(file_path)
            wb.close()
            
            return True
            
        except Exception as e:
            print(f"❌ Error writing XLSX: {e}")
            return False
    
    def _write_json(self, sheets_data: Dict[str, pd.DataFrame], 
                    file_path: str, metadata: WorkbookMetadata = None) -> bool:
        """Write to JSON format with configurable structure."""
        try:
            output_data = {}
            
            # Handle different JSON orientations
            if self.config.combine_sheets and len(sheets_data) > 1:
                # Combine all sheets
                combined_data = []
                for sheet_name, df in sheets_data.items():
                    sheet_json = df.to_dict(orient=self.config.json_orient)
                    if self.config.json_orient == 'records':
                        # Add sheet name to each record
                        for record in sheet_json:
                            record['_sheet_name'] = sheet_name
                        combined_data.extend(sheet_json)
                    else:
                        combined_data.append({sheet_name: sheet_json})
                
                output_data = combined_data
            else:
                # Separate sheets
                for sheet_name, df in sheets_data.items():
                    output_data[sheet_name] = df.to_dict(orient=self.config.json_orient)
            
            # Add metadata if requested
            if metadata and self.config.json_include_metadata:
                output_data['_metadata'] = asdict(metadata)
            
            # Write JSON file
            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(output_data, f, indent=self.config.json_indent, default=str, ensure_ascii=False)
            
            return True
            
        except Exception as e:
            print(f"❌ Error writing JSON: {e}")
            return False
    
    def _write_yaml(self, sheets_data: Dict[str, pd.DataFrame], 
                    file_path: str, metadata: WorkbookMetadata = None) -> bool:
        """Write to YAML format."""
        if not yaml:
            print("❌ Error: PyYAML not installed. Cannot write YAML format.")
            return False
            
        try:
            output_data = {}
            
            # Convert DataFrames to YAML-friendly format
            if self.config.combine_sheets and len(sheets_data) > 1:
                # Combine all sheets
                combined_data = []
                for sheet_name, df in sheets_data.items():
                    sheet_data = df.to_dict(orient='records')
                    # Add sheet identifier
                    for record in sheet_data:
                        record['_sheet_name'] = sheet_name
                    combined_data.extend(sheet_data)
                
                output_data = {'data': combined_data}
            else:
                # Separate sheets
                for sheet_name, df in sheets_data.items():
                    output_data[sheet_name] = df.to_dict(orient='records')
            
            # Add metadata if requested
            if metadata and self.config.yaml_include_metadata:
                output_data['_metadata'] = asdict(metadata)
            
            # Write YAML file
            with open(file_path, 'w', encoding='utf-8') as f:
                yaml.dump(
                    output_data, 
                    f, 
                    default_flow_style=self.config.yaml_default_flow_style,
                    allow_unicode=True,
                    sort_keys=False
                )
            
            return True
            
        except Exception as e:
            print(f"❌ Error writing YAML: {e}")
            return False
    
    def _write_markdown(self, sheets_data: Dict[str, pd.DataFrame], 
                        file_path: str, metadata: WorkbookMetadata = None) -> bool:
        """Write to Markdown table format."""
        try:
            markdown_lines = []
            
            # Add title if multiple sheets
            if len(sheets_data) > 1:
                markdown_lines.append("# Spreadsheet Data\n")
            
            for sheet_name, df in sheets_data.items():
                # Add sheet heading if multiple sheets or requested
                if len(sheets_data) > 1 or self.config.md_include_sheet_names:
                    markdown_lines.append(f"## {sheet_name}\n")
                
                if df.empty:
                    markdown_lines.append("*No data in this sheet*\n")
                    continue
                
                # Truncate long content for readability
                display_df = df.copy()
                for col in display_df.select_dtypes(include=['object']):
                    display_df[col] = display_df[col].astype(str).apply(
                        lambda x: x[:self.config.md_max_col_width] + '...' 
                        if len(x) > self.config.md_max_col_width else x
                    )
                
                # Convert DataFrame to markdown table
                md_table = display_df.to_markdown(
                    index=False,
                    tablefmt='github',
                    stralign=self.config.md_table_alignment
                )
                
                if md_table:
                    markdown_lines.append(md_table)
                    markdown_lines.append("\n")
            
            # Add metadata if requested
            if metadata and self.config.yaml_include_metadata:
                markdown_lines.append("## Metadata\n")
                markdown_lines.append("```yaml")
                markdown_lines.append(yaml.dump(asdict(metadata), default_flow_style=False) if yaml else str(metadata))
                markdown_lines.append("```\n")
            
            # Write markdown file
            with open(file_path, 'w', encoding='utf-8') as f:
                f.write('\n'.join(markdown_lines))
            
            return True
            
        except Exception as e:
            print(f"❌ Error writing Markdown: {e}")
            return False


class ExcelConverter:
    """Main converter class that orchestrates reading and writing."""
    
    def __init__(self, config: ConversionConfig = None, safety_manager: FileSafetyManager = None):
        self.config = config or ConversionConfig()
        self.safety = safety_manager or FileSafetyManager()
        self.reader = ExcelReader(self.config)
        self.writer = ExcelWriter(self.config)
    
    def convert_file(self, input_path: str, output_path: str) -> bool:
        """
        Convert file from one format to another.
        Returns: success status
        """
        source_file = Path(input_path)
        target_file = Path(output_path)
        
        if not source_file.exists():
            print(f"❌ Error: Source file {source_file} does not exist")
            return False
        
        # Safety check
        can_proceed, reason = self.safety.safe_write_check(source_file, target_file)
        if not can_proceed:
            print(f"❌ Safety check failed: {reason}")
            return False
        
        try:
            # Read source file
            sheets_data, metadata = self.reader.read_file(input_path)
            
            # Add conversion metadata
            metadata.file_hash = self.safety.calculate_file_hash(source_file)
            metadata.conversion_timestamp = datetime.now().isoformat()
            
            # Write to target format
            success = self.writer.write_file(sheets_data, output_path, metadata)
            
            if success:
                output_hash = self.safety.calculate_file_hash(target_file)
                print(f"✅ Successfully converted {source_file} → {target_file}")
                print(f"📊 Output hash: {output_hash[:16]}...")
                return True
            else:
                print(f"❌ Conversion failed")
                return False
            
        except Exception as e:
            print(f"❌ Error converting {source_file}: {e}")
            return False


def load_config_file(config_path: str) -> ConversionConfig:
    """Load configuration from file."""
    if not Path(config_path).exists():
        print(f"⚠️  Config file {config_path} not found, using defaults")
        return ConversionConfig()
    
    try:
        with open(config_path, 'r', encoding='utf-8') as f:
            if config_path.endswith('.json'):
                config_dict = json.load(f)
            elif config_path.endswith(('.yml', '.yaml')) and yaml:
                config_dict = yaml.safe_load(f)
            else:
                print("⚠️  Unsupported config format, using defaults")
                return ConversionConfig()
        
        config = ConversionConfig()
        for key, value in config_dict.items():
            if hasattr(config, key):
                setattr(config, key, value)
        
        return config
    except Exception as e:
        print(f"⚠️  Error loading config: {e}, using defaults")
        return ConversionConfig()


def create_sample_config(config_path: str) -> None:
    """Create a sample configuration file."""
    config = ConversionConfig()
    config_dict = asdict(config)
    
    try:
        if config_path.endswith('.json'):
            with open(config_path, 'w', encoding='utf-8') as f:
                json.dump(config_dict, f, indent=2)
        elif config_path.endswith(('.yml', '.yaml')) and yaml:
            with open(config_path, 'w', encoding='utf-8') as f:
                yaml.dump(config_dict, f, default_flow_style=False)
        else:
            print("❌ Unsupported config format")
            return
        
        print(f"✅ Sample configuration created at {config_path}")
    except Exception as e:
        print(f"❌ Error creating config: {e}")


def main():
    """Main function to handle command line arguments with safety features."""
    parser = argparse.ArgumentParser(
        description='Excel - Safe spreadsheet conversion tool with hash validation',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  excel data.xlsx data.csv                # Convert Excel to CSV
  excel data.csv data.xlsx                # Convert CSV to Excel
  excel data.xlsx data.json               # Convert Excel to JSON
  excel data.xlsx data.md                 # Convert Excel to Markdown tables
  excel --sheet "Sheet1,Sheet2" data.xlsx output.csv  # Convert specific sheets
  excel --combine-sheets data.xlsx combined.csv       # Combine all sheets
  excel --create-config config.yaml       # Create sample configuration

Supported Formats:
  Input:  .xlsx, .xlsm, .xls, .csv, .tsv
  Output: .xlsx, .xlsm, .csv, .tsv, .json, .yaml, .md

Safety Features:
  - Hash validation prevents data loss
  - Automatic backup creation
  - Collision detection  
  - Confirmation prompts for overwrites
  - Metadata preservation for round-trip conversions
        """
    )
    
    parser.add_argument('input_file', nargs='?', help='Input file path')
    parser.add_argument('output_file', nargs='?', help='Output file path')
    parser.add_argument('--config', help='Configuration file path (JSON or YAML)')
    parser.add_argument('--create-config', help='Create sample configuration file')
    
    # Sheet selection
    parser.add_argument('--sheet', help='Comma-separated list of sheet names to convert')
    parser.add_argument('--combine-sheets', action='store_true', 
                       help='Combine all sheets into single output')
    
    # Format options
    parser.add_argument('--delimiter', default=',', help='CSV delimiter (default: comma)')
    parser.add_argument('--encoding', default='utf-8', help='File encoding (default: utf-8)')
    parser.add_argument('--no-header', action='store_true', help='Skip header row')
    parser.add_argument('--no-index', action='store_true', help='Skip row index')
    
    # Processing options
    parser.add_argument('--skip-empty', action='store_true', help='Skip empty rows/columns')
    parser.add_argument('--no-formatting', action='store_true', help='Skip formatting preservation')
    parser.add_argument('--no-formulas', action='store_true', help='Skip formula preservation')
    
    # Safety options
    parser.add_argument('--force', action='store_true', help='Skip confirmation prompts')
    parser.add_argument('--no-backup', action='store_true', help='Skip backup creation')
    parser.add_argument('--no-hash-check', action='store_true', help='Skip hash validation')
    
    args = parser.parse_args()
    
    # Handle config creation
    if args.create_config:
        create_sample_config(args.create_config)
        return
    
    if not args.input_file or not args.output_file:
        parser.print_help()
        sys.exit(1)
    
    input_path = Path(args.input_file)
    output_path = Path(args.output_file)
    
    if not input_path.exists():
        print(f"❌ Error: Input file {input_path} does not exist")
        sys.exit(1)
    
    # Configure safety settings
    safety_config = SafetyConfig(
        require_confirmation=not args.force,
        create_backup=not args.no_backup,
        check_hash=not args.no_hash_check,
        prevent_overwrite=True
    )
    
    safety_manager = FileSafetyManager(safety_config)
    
    # Load and configure conversion settings
    config = load_config_file(args.config) if args.config else ConversionConfig()
    
    # Apply command line overrides
    if args.sheet:
        config.sheet_selection = [s.strip() for s in args.sheet.split(',')]
    if args.combine_sheets:
        config.combine_sheets = True
    if args.delimiter:
        config.csv_delimiter = args.delimiter
    if args.encoding:
        config.csv_encoding = args.encoding
    if args.no_header:
        config.csv_include_header = False
        config.excel_header = False
    if args.no_index:
        config.csv_include_index = False
        config.excel_index = False
    if args.skip_empty:
        config.skip_empty_rows = True
        config.skip_empty_columns = True
    if args.no_formatting:
        config.preserve_formatting = False
    if args.no_formulas:
        config.preserve_formulas = False
    
    try:
        converter = ExcelConverter(config, safety_manager)
        success = converter.convert_file(str(input_path), str(output_path))
        sys.exit(0 if success else 1)
            
    except Exception as e:
        print(f"❌ Conversion failed: {e}")
        sys.exit(1)


if __name__ == '__main__':
    main()