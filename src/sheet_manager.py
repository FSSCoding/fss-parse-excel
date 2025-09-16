#!/usr/bin/env python3
"""
Sheet Manager - Sheet-level operations for Excel files
Handles sheet creation, deletion, copying, and management.
"""

from pathlib import Path
from typing import List, Optional
from openpyxl import load_workbook, Workbook
from converters import SafetyConfig, FileSafetyManager


class SheetManager:
    """
    Manages sheet-level operations with safety controls.
    """
    
    def __init__(self, file_path: Path, safety_config: SafetyConfig):
        self.file_path = file_path
        self.safety_config = safety_config
        self.safety_manager = FileSafetyManager(safety_config)
    
    def list_sheets(self) -> List[str]:
        """List all sheet names."""
        try:
            workbook = load_workbook(self.file_path)
            return workbook.sheetnames
        except Exception as e:
            return []
    
    def add_sheet(self, name: str, template: str = None) -> bool:
        """Add a new sheet."""
        try:
            workbook = load_workbook(self.file_path)
            if name in workbook.sheetnames:
                return False
            workbook.create_sheet(name)
            workbook.save(self.file_path)
            return True
        except Exception:
            return False
    
    def delete_sheet(self, name: str) -> bool:
        """Delete a sheet."""
        try:
            workbook = load_workbook(self.file_path)
            if name not in workbook.sheetnames or len(workbook.sheetnames) <= 1:
                return False
            workbook.remove(workbook[name])
            workbook.save(self.file_path)
            return True
        except Exception:
            return False